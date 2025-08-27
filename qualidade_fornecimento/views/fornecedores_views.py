# qualidade_fornecimento/views/fornecedores_views.py
from datetime import timedelta

import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from qualidade_fornecimento.forms import FornecedorForm
from qualidade_fornecimento.models import FornecedorQualificado


from datetime import timedelta

from django.shortcuts import render
from django.core.paginator import Paginator
from django.utils import timezone
from django.contrib.auth.decorators import login_required, permission_required

from qualidade_fornecimento.models import FornecedorQualificado

@login_required
@permission_required('qualidade_fornecimento.view_fornecedorqualificado', raise_exception=True)
def lista_fornecedores(request):
    # Base para montar os selects
    fornecedores_base_qs = FornecedorQualificado.objects.all()

    # Queryset principal, ordenado por nome
    fornecedores_qs = fornecedores_base_qs.order_by("nome")
    filter_nomes = sorted(set(fornecedores_base_qs.values_list("nome", flat=True)))

    # Captura dos filtros via GET
    nome = request.GET.get("nome", "").strip()                # ← novo filtro
    data_inicial = request.GET.get("data_inicial")
    data_final   = request.GET.get("data_final")
    produto      = request.GET.get("produto")
    certificacao = request.GET.get("certificacao")
    ativo_filter = request.GET.get("status")  # "Ativo", "Inativo" ou None

    # Aplicação dos filtros em cascata
    if nome:
        fornecedores_qs = fornecedores_qs.filter(nome__icontains=nome)
    if data_inicial:
        fornecedores_qs = fornecedores_qs.filter(data_homologacao__gte=data_inicial)
    if data_final:
        fornecedores_qs = fornecedores_qs.filter(data_homologacao__lte=data_final)
    if produto:
        fornecedores_qs = fornecedores_qs.filter(produto_servico=produto)
    if certificacao:
        fornecedores_qs = fornecedores_qs.filter(tipo_certificacao=certificacao)

    # Filtro padrão de status: se não veio nenhum, mostra apenas "Ativo"
    if ativo_filter:
        fornecedores_qs = fornecedores_qs.filter(ativo=ativo_filter)
    else:
        fornecedores_qs = fornecedores_qs.filter(ativo__in=["Ativo", "Em Homologação"])

    # Dados para popular os selects (exceto nome, que agora é input text)
    filter_produtos      = sorted(set(fornecedores_base_qs.values_list("produto_servico", flat=True)))
    filter_certificacoes = sorted(set(fornecedores_base_qs.values_list("tipo_certificacao", flat=True)))
    filter_status        = ["Ativo", "Em Homologação", "Inativo"]

    # Paginação
    paginator  = Paginator(fornecedores_qs, 10)
    page_obj   = paginator.get_page(request.GET.get("page"))

    # Cálculo de indicadores
    hoje           = timezone.now().date()
    ate_30_dias    = hoje + timedelta(days=30)
    total_forn     = fornecedores_qs.count()
    total_vencidas = fornecedores_qs.filter(vencimento_certificacao__lt=hoje).count()
    total_prox     = fornecedores_qs.filter(
        vencimento_certificacao__gte=hoje,
        vencimento_certificacao__lte=ate_30_dias,
    ).count()
    total_risco_alto = fornecedores_qs.filter(risco="Alto").count()

    # Monta contexto e renderiza template
    context = {
        "fornecedores_paginados": page_obj,
        "total_fornecedores":     total_forn,
        "total_vencidas":         total_vencidas,
        "total_proximas":         total_prox,
        "total_alto_risco":       total_risco_alto,
        "current_date":           hoje,
        "current_date_plus_30":   ate_30_dias,
        "filter_produtos":        filter_produtos,
        "filter_certificacoes":   filter_certificacoes,
        "filter_status":          filter_status,
        "filter_nomes": filter_nomes,

    }
    return render(request, "fornecedores/lista_fornecedores.html", context)



import logging

from django.contrib import messages

logger = logging.getLogger(__name__)


@login_required
def cadastrar_fornecedor(request):
    if request.method == "POST":
        form = FornecedorForm(request.POST, request.FILES)
        if form.is_valid():
            obj = form.save(commit=False)
            if obj.vencimento_certificacao:
                obj.data_avaliacao_risco = obj.vencimento_certificacao
            obj.save()
            try:
                form.save()
                messages.success(request, "Fornecedor cadastrado com sucesso!")
                return redirect("lista_fornecedores")
            except Exception as e:
                logger.error("Erro ao salvar o fornecedor: %s", e, exc_info=True)
                messages.error(request, f"Erro ao salvar o fornecedor: {e}")
        else:
            error_messages = []
            for field, errors in form.errors.items():
                error_messages.append(f"{field}: {', '.join(errors)}")
            messages.error(
                request, "Erro ao cadastrar o fornecedor. " + " | ".join(error_messages)
            )
    else:
        form = FornecedorForm()

    return render(
        request, "fornecedores/form_fornecedor.html", {"form": form, "modo": "Cadastro"}
    )


@login_required
@permission_required('qualidade_fornecimento.change_fornecedorqualificado', raise_exception=True)
def editar_fornecedor(request, id):
    fornecedor = get_object_or_404(FornecedorQualificado, id=id)

    if request.method == "POST":
        form = FornecedorForm(request.POST, request.FILES, instance=fornecedor)
        if form.is_valid():
            try:
                # Salva sem confirmar no DB para permitir tratar o arquivo antes
                obj = form.save(commit=False)
                if obj.vencimento_certificacao:
                    obj.data_avaliacao_risco = obj.vencimento_certificacao
                obj.save()

                # Suporte a dois jeitos de pedir remoção do arquivo:
                # 1) Nosso checkbox custom: remover_certificado=1
                # 2) Checkbox padrão do Django ClearableFileInput: certificado_anexo-clear=on
                remove_custom = request.POST.get("remover_certificado") == "1"
                remove_default = request.POST.get("certificado_anexo-clear") == "on"
                if (remove_custom or remove_default) and fornecedor.certificado_anexo:
                    # Apaga o arquivo físico e limpa o campo
                    fornecedor.certificado_anexo.delete(save=False)
                    obj.certificado_anexo = None

                # Obs: se veio um novo arquivo em request.FILES, o form já atualizou em obj

                obj.save()
                # (não há M2M nesse form; se houvesse, chamaríamos form.save_m2m())

                messages.success(request, "Fornecedor atualizado com sucesso!")
                return redirect("lista_fornecedores")

            except Exception as e:
                logger.error("Erro ao salvar o fornecedor: %s", e, exc_info=True)
                messages.error(request, f"Erro ao salvar o fornecedor: {e}")
        else:
            error_messages = []
            for field, errors in form.errors.items():
                error_messages.append(f"{field}: {', '.join(errors)}")
            messages.error(
                request,
                "Erro ao atualizar o fornecedor. " + " | ".join(error_messages),
            )
    else:
        form = FornecedorForm(instance=fornecedor)

    return render(
        request,
        "fornecedores/form_fornecedor.html",
        {"form": form, "modo": "Edição"},
    )


@login_required
def excluir_fornecedor(request, id):
    fornecedor = get_object_or_404(FornecedorQualificado, id=id)
    if request.method == "POST":
        fornecedor.delete()
        messages.success(request, "Fornecedor excluído com sucesso!")
        return redirect("lista_fornecedores")
    return render(
        request, "fornecedores/excluir_fornecedor.html", {"fornecedor": fornecedor}
    )


@login_required
def visualizar_fornecedor(request, id):
    fornecedor = get_object_or_404(FornecedorQualificado, id=id)
    context = {"fornecedor": fornecedor, "now": timezone.now()}
    return render(request, "fornecedores/visualizar_fornecedor_pdf.html", context)



@login_required
def importar_excel_fornecedores(request):
    if request.method == "POST":
        excel_file = request.FILES.get("excel_file")
        if excel_file:
            try:
                workbook = openpyxl.load_workbook(excel_file)
                sheet = workbook.active
                row_count = sheet.max_row
                logger.info("Total de linhas no Excel (incluindo cabeçalho): %s", row_count)

                erros = []  # Lista para capturar os erros por linha

                for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        # Ignora linhas totalmente vazias
                        if not any(row):
                            logger.info("Linha %s vazia, pulando", index)
                            continue

                        nome = row[0]
                        produto_servico = row[1]
                        data_homologacao = row[2]
                        tipo_certificacao = row[3]
                        vencimento_certificacao = row[4] or None
                        risco = row[5] or None
                        data_avaliacao_risco = row[6] or None
                        tipo_formulario = row[7] or None
                        data_auditoria = row[8] or None
                        raw_nota_auditoria = row[9]
                        try:
                            nota_auditoria = float(raw_nota_auditoria) if raw_nota_auditoria is not None else None
                        except ValueError:
                            nota_auditoria = None                        
                        especialista_nome = row[10] or "Não informado"
                        especialista_contato = row[11] or "Não informado"
                        especialista_cargo = row[12] or "Não informado"
                        fornecedor, created = FornecedorQualificado.objects.get_or_create(
                            nome=nome,
                            defaults={
                                "produto_servico": produto_servico,
                                "data_homologacao": data_homologacao,
                                "tipo_certificacao": tipo_certificacao,
                                "vencimento_certificacao": vencimento_certificacao,
                                "risco": risco,
                                "data_avaliacao_risco": data_avaliacao_risco,
                                "tipo_formulario": tipo_formulario,
                                "data_auditoria": data_auditoria,
                                "nota_auditoria": nota_auditoria,
                                "especialista_nome": especialista_nome,
                                "especialista_contato": especialista_contato,
                                "especialista_cargo": especialista_cargo,
                            },
                        )
                        if not created:
                            fornecedor.produto_servico = produto_servico
                            fornecedor.data_homologacao = data_homologacao
                            fornecedor.tipo_certificacao = tipo_certificacao
                            fornecedor.vencimento_certificacao = vencimento_certificacao
                            fornecedor.risco = risco
                            fornecedor.data_avaliacao_risco = data_avaliacao_risco
                            fornecedor.tipo_formulario = tipo_formulario
                            fornecedor.data_auditoria = data_auditoria
                            fornecedor.nota_auditoria = nota_auditoria
                            fornecedor.especialista_nome = especialista_nome
                            fornecedor.especialista_contato = especialista_contato
                            fornecedor.especialista_cargo = especialista_cargo
                            fornecedor.save()

                    except Exception as linha_erro:
                        mensagem_erro = f"Linha {index}: erro ao importar - {str(linha_erro)}"
                        logger.error(mensagem_erro)
                        erros.append(mensagem_erro)

                if erros:
                    messages.warning(request, "Algumas linhas não foram importadas corretamente.")
                    for erro in erros:
                        messages.warning(request, erro)
                else:
                    messages.success(request, "Todos os dados foram importados com sucesso!")

                return redirect("lista_fornecedores")

            except Exception as e:
                logger.error("Erro ao importar arquivo: %s", e, exc_info=True)
                messages.error(request, f"Erro ao importar arquivo: {e}")
        else:
            messages.error(request, "Selecione um arquivo Excel para importar.")
    return render(request, "fornecedores/importar_excel.html")

