# qualidade_fornecimento/views/materiaprima_views.py

import base64
import logging
from datetime import date
from django.templatetags.static import static
from django.conf import settings
import os

import openpyxl  # se desejar usar a importação via Excel
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone

from qualidade_fornecimento.forms.tb050_forms import RelacaoMateriaPrimaForm
from qualidade_fornecimento.models.f045 import RelatorioF045
from qualidade_fornecimento.models.rolo import RoloMateriaPrima

logger = logging.getLogger(__name__)

from django.forms import inlineformset_factory
from django.shortcuts import render

from qualidade_fornecimento.forms.inline_rolo_formset import RoloFormSet
from qualidade_fornecimento.models import (
    FornecedorQualificado,
    MateriaPrimaCatalogo,
    RelacaoMateriaPrima,
)
from qualidade_fornecimento.models.rolo import RoloMateriaPrima


@login_required
def lista_tb050(request):
    """
    Lista/pesquisa de registros TB050 com filtros, indicadores,
    paginação e queryset adicional (relacoes) para o modal “Gerar F045”.
    """
    # ------------------ 1. Base Queryset ------------------
    qs = RelacaoMateriaPrima.objects.all().order_by("-atualizado_em")

    # ------------------ 2. Filtros GET --------------------
    data_inicial = request.GET.get("data_inicial")
    data_final = request.GET.get("data_final")
    status_filter = request.GET.get("status")
    fornecedor_filter = request.GET.get("fornecedor")
    materia_prima_filter = request.GET.get("materia_prima")
    nro_rolo = request.GET.get("nro_rolo")
    nro_relatorio = request.GET.get("nro_relatorio")
    nro_certificado = request.GET.get("nro_certificado")

    if data_inicial:
        qs = qs.filter(data_entrada__gte=data_inicial)
    if data_final:
        qs = qs.filter(data_entrada__lte=data_final)
    if status_filter:
        qs = qs.filter(status=status_filter)
    if fornecedor_filter:
        qs = qs.filter(fornecedor_id=fornecedor_filter)
    if materia_prima_filter:
        qs = qs.filter(materia_prima_id=materia_prima_filter)
    if nro_rolo:
        qs = qs.filter(rolos__nro_rolo__icontains=nro_rolo).distinct()
    if nro_relatorio:
        qs = qs.filter(nro_relatorio__icontains=nro_relatorio)
    if nro_certificado:
        qs = qs.filter(numero_certificado__icontains=nro_certificado)

    # ------------------ 3. Paginação ----------------------
    paginator = Paginator(qs, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # ------------------ 4. Indicadores --------------------
    total_registros = qs.count()
    total_aprovados = qs.filter(status="Aprovado").count()
    total_reprovados = qs.filter(status="Reprovado").count()
    total_atrasados = qs.filter(atraso_em_dias__gt=0).count()

    # ------------------ 5. Listas dinâmicas ---------------
    lista_fornecedores = FornecedorQualificado.objects.filter(
        id__in=qs.values_list("fornecedor_id", flat=True)
    ).order_by("nome")

    lista_materiasprimas = MateriaPrimaCatalogo.objects.filter(
        id__in=qs.values_list("materia_prima_id", flat=True)
    ).order_by("codigo")

    # ------------------ 6. Queryset para o modal F045 ----
    #  • Usa o mesmo filtro aplicado na lista
    #  • Adiciona select_related para evitar N+1
    #  • Descomente a linha com f045__isnull=True se quiser mostrar
    #    somente relatórios que ainda não têm F045 gerado.
    relacoes = (
        qs.select_related("materia_prima").order_by("-nro_relatorio")
        # .filter(f045__isnull=True)   # opcional
    )

    f045_pending = request.session.pop("f045_pending", None)

    # ------------------ 7. Contexto -----------------------
    context = {
        "materias_primas_paginadas": page_obj,
        "filter_status": sorted(
            set(
                s
                for s in RelacaoMateriaPrima.objects.values_list("status", flat=True)
                if s is not None
            )
        ),
        "total_registros": total_registros,
        "total_aprovados": total_aprovados,
        "total_reprovados": total_reprovados,
        "total_atrasados": total_atrasados,
        "data_inicial": data_inicial,
        "data_final": data_final,
        "status_filter": status_filter,
        "lista_fornecedores": lista_fornecedores,
        "lista_materiasprimas": lista_materiasprimas,
        "relacoes": relacoes,  # para o modal
        "f045_pending": f045_pending,
    }

    return render(request, "tb050/lista_tb050.html", context)


def salvar_formulario_tb050(form, formset, registro_existente=None):
    tb050 = form.save(commit=False)

    if tb050.nro_relatorio:
        try:
            from qualidade_fornecimento.models.f045 import RelatorioF045

            f045 = RelatorioF045.objects.get(nro_relatorio=tb050.nro_relatorio)
            status = f045.status_geral.strip().lower()

            if status == "aprovado":
                tb050.status = "Aprovado"
            elif status in [
                "aprovado condicional",
                "aprovação condicional",
                "aprovado condicionalmente",
            ]:
                tb050.status = "Aprovado Condicionalmente"
            elif status == "reprovado":
                tb050.status = "Reprovado"
            else:
                tb050.status = "Aguardando F045"
        except RelatorioF045.DoesNotExist:
            tb050.status = "Aguardando F045"
    else:
        tb050.status = "Aguardando F045"

    tb050.save()

    # 1️⃣ Salva primeiro os objetos em memória
    rolos = formset.save(commit=False)

    # 2️⃣ Agora sim: Deleta os que marcaram como excluir
    if hasattr(formset, "deleted_forms"):
        for form in formset.deleted_forms:
            if form.instance.pk:
                form.instance.delete()

    # 3️⃣ Depois salva os novos ou editados
    # … dentro de salvar_formulario_tb050 …
    for rolo in rolos:
        if not rolo.nro_rolo or rolo.nro_rolo.strip() in ["", "Será gerado ao salvar"]:
            # ✅ filtra apenas os rolos desta relação
            qs = RoloMateriaPrima.objects.filter(tb050=tb050)
            # extrai apenas strings numéricas e converte em int
            numeros = [
                int(n) for n in qs.values_list("nro_rolo", flat=True)
                if n and n.isdigit()
            ]
            # se não houver nenhum, começa em 50000
            ultimo_numero = max(numeros) if numeros else 49999
            rolo.nro_rolo = str(ultimo_numero + 1)

        # garante FK antes de salvar qualquer rolo (novo ou editado)
        rolo.tb050 = tb050
        rolo.save()


    formset.save_m2m()

    return tb050


@login_required
def cadastrar_tb050(request):
    if request.method == "POST":
        form = RelacaoMateriaPrimaForm(request.POST, request.FILES)
        if form.is_valid():
            # Assegura que novo registro já venha com status "Aguardando F045"
            tb050 = form.save(commit=False)
            tb050.status = "Aguardando F045"
            tb050.save()
            messages.success(request, "Registro da TB050 cadastrado com sucesso!")
            return redirect("tb050_list")
    else:
        form = RelacaoMateriaPrimaForm()

    return render(
        request,
        "tb050/form_tb050.html",
        {"form": form, "titulo_pagina": "Cadastrar Relação de Matérias-Primas (TB050)"},
    )



from qualidade_fornecimento.forms.inline_rolo_formset import RoloFormSet

@login_required
def editar_tb050(request, id):
    registro = get_object_or_404(RelacaoMateriaPrima, id=id)

    if request.method == "POST":
        form = RelacaoMateriaPrimaForm(request.POST, request.FILES, instance=registro)
        if form.is_valid():
            tb050 = form.save(commit=False)
            try:
                tb050.status = tb050.f045.status_geral
            except RelatorioF045.DoesNotExist:
                tb050.status = "Aguardando F045"
            tb050.save()
            messages.success(request, "Registro atualizado com sucesso!")
            return redirect("tb050_list")
    else:
        form = RelacaoMateriaPrimaForm(instance=registro)

    return render(request, "tb050/form_tb050.html", {
        "form": form,
        "registro": registro,
        "titulo_pagina": "Editar Relação de Matérias-Primas (TB050)",
    })




@login_required
def excluir_tb050(request, id):
    registro = get_object_or_404(RelacaoMateriaPrima, id=id)
    if request.method == "POST":
        registro.delete()
        messages.success(request, "Relatório excluído com sucesso.")
    return redirect("tb050_list")


@login_required
def visualizar_tb050(request, id):
    """
    Visualiza os detalhes de um registro da TB050 – Relação de Matérias-Primas.
    """
    registro = get_object_or_404(RelacaoMateriaPrima, id=id)
    context = {
        "registro": registro,
        "now": timezone.now(),
    }
    return render(request, "tb050/visualizar_tb050.html", context)


@login_required
def importar_excel_tb050(request):
    if request.method == "POST":
        excel_file = request.FILES.get("excel_file")
        if not excel_file:
            messages.error(request, "Selecione um arquivo Excel para importar.")
            return redirect("tb050_importar_excel")

        try:
            workbook = openpyxl.load_workbook(excel_file)
            sheet = workbook.active
            row_count = sheet.max_row
            logger.info("Total de linhas no Excel: %s", row_count)

            for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue

                data_entrada = row[0]
                fornecedor_nome = row[1]
                nota_fiscal = row[2]
                numero_certificado = row[3]
                materia_prima_codigo = row[4]
                data_prevista_entrega = row[5]
                data_renegociada_entrega = row[6]

                fornecedor = FornecedorQualificado.objects.filter(
                    nome__iexact=fornecedor_nome.strip()
                ).first()
                if not fornecedor:
                    messages.warning(request, f"Linha {index}: Fornecedor '{fornecedor_nome}' não encontrado.")
                    continue

                materia_prima = MateriaPrimaCatalogo.objects.filter(
                    codigo__iexact=materia_prima_codigo.strip()
                ).first()
                if not materia_prima:
                    messages.warning(request, f"Linha {index}: Matéria-prima com código '{materia_prima_codigo}' não encontrada.")
                    continue

                RelacaoMateriaPrima.objects.create(
                    data_entrada=data_entrada,
                    fornecedor=fornecedor,
                    nota_fiscal=nota_fiscal,
                    numero_certificado=numero_certificado,
                    materia_prima=materia_prima,
                    data_prevista_entrega=data_prevista_entrega,
                    data_renegociada_entrega=data_renegociada_entrega,
                )

            messages.success(request, "Importação concluída com sucesso.")
            return redirect("tb050_list")

        except Exception as e:
            logger.error("Erro ao importar TB050: %s", e, exc_info=True)
            messages.error(request, f"Erro ao importar: {e}")

    return render(request, "tb050/importar_excel_tb050.html")


import io

from django.http import HttpResponse, JsonResponse
from django.template.loader import get_template
from xhtml2pdf import pisa


@login_required
def selecionar_etiquetas_tb050(request, id):
    registro = get_object_or_404(RelacaoMateriaPrima, id=id)
    rolos = registro.rolos.all()

    # Debug: Verifique se os valores estão sendo carregados
    for rolo in rolos:
        print(f"Rolo ID: {rolo.nro_rolo}, Peso: {rolo.peso}")

    return render(
        request,
        "tb050/selecionar_etiquetas.html",
        {
            "registro": registro,
            "rolos": rolos,
        },
    )


@login_required
def imprimir_etiquetas_tb050(request, id):
    registro = get_object_or_404(RelacaoMateriaPrima, id=id)

    if request.method == "POST":
        # Pega os IDs dos rolos selecionados
        rolos_ids = request.POST.getlist("rolos")
        rolos_selecionados = []

        for rolo_id in rolos_ids:
            # Busca pelo campo correto
            rolo = RoloMateriaPrima.objects.get(nro_rolo=rolo_id)

            # Campos do POST usando rolo.nro_rolo
            peso_str = request.POST.get(f"peso_{rolo.nro_rolo}")
            qtd_str  = request.POST.get(f"quantidade_{rolo.nro_rolo}")


            # Atualiza os valores no rolo (sem salvar no banco, apenas em memória)
            rolo.peso = float(peso_str.replace(",", ".")) if peso_str else 0.0
            rolo.qtd_etiquetas = int(qtd_str) if qtd_str else 1

            rolos_selecionados.append(rolo)

        # Gera QR Code e cria o PDF (direto, sem preview)
        for rolo in rolos_selecionados:
            if rolo.peso is None:
                rolo.peso = 0.0
            data_qr = (
                f"{registro.materia_prima.codigo};"
                f"{rolo.nro_rolo};"
                f"{rolo.peso};"
                f"{registro.materia_prima.localizacao};"
                f"{registro.fornecedor.nome}"
            )
            qr = qrcode.make(data_qr)
            buffer = io.BytesIO()
            qr.save(buffer, format="PNG")
            qr_code_base64 = base64.b64encode(buffer.getvalue()).decode()
            rolo.qrcode_url = f"data:image/png;base64,{qr_code_base64}"

        # Renderiza o template PDF diretamente
        logo_url = f"file://{os.path.join(settings.STATIC_ROOT, 'img', 'logo.png')}"
        seguranca_url = f"file://{os.path.join(settings.STATIC_ROOT, 'img', 'seguranca.png')}"

        html_string = render_to_string(
            "tb050/etiqueta_lote_pdf.html",
            {
                "registro": registro,
                "rolos": rolos_selecionados,
                "request": request,
                "data_atual": date.today().strftime("%d/%m/%Y"),
                "logo_url": logo_url,
                "seguranca_url": seguranca_url,  # ✅ novo
            },
        )


        with tempfile.NamedTemporaryFile(delete=True) as output:
            html = HTML(string=html_string, base_url=request.build_absolute_uri("/"))
            pdf_bytes = html.write_pdf()
            output.write(pdf_bytes)
            output.seek(0)
            response = HttpResponse(output.read(), content_type="application/pdf")
            response["Content-Disposition"] = (
                f"filename=etiquetas_lote_{registro.id}.pdf"
            )
            return response

    # Se não for POST (ou seja, GET), redireciona para a página de seleção
    return redirect("tb050_selecionar_etiquetas", id=id)


def get_rolos_peso(request, id):
    registro = get_object_or_404(RelacaoMateriaPrima, id=id)
    rolos = registro.rolos.all()

    data = {
        "rolos": [
            {
                "id": rolo.nro_rolo,
                "nro_rolo": rolo.nro_rolo,
                "peso": float(rolo.peso) if rolo.peso is not None else 0.0,
            }
            for rolo in rolos
        ]
    }


    return JsonResponse(data)


import tempfile
from datetime import datetime
from io import BytesIO

import qrcode
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.template.loader import get_template, render_to_string
from weasyprint import CSS, HTML
from xhtml2pdf import pisa


@login_required
def imprimir_etiquetas_pdf(request, id):
    registro = get_object_or_404(RelacaoMateriaPrima, id=id)
    # Inicialmente, pega todos os rolos relacionados
    rolos_queryset = registro.rolos.all()

    if request.method == "POST":
        # Se houver algum filtro enviado via campo oculto, aplica-o.
        # Exemplo: filtro pelo número do rolo
        nro_rolo_filter = request.POST.get("nro_rolo")
        if nro_rolo_filter:
            rolos_queryset = rolos_queryset.filter(nro_rolo__icontains=nro_rolo_filter)

        # Agora, pega os rolos que foram selecionados via checkbox
        rolos_ids = request.POST.getlist("rolos")
        if rolos_ids:
            rolos_queryset = rolos_queryset.filter(nro_rolo__in=rolos_ids)
        # Caso não haja rolos selecionados explicitamente, usa todos os que passaram nos filtros

        # Atualiza os valores de peso e quantidade conforme o formulário
        rolos = []
        for rolo in rolos:
            peso_str = request.POST.get(f"peso_{rolo.nro_rolo}")
            qtd_str  = request.POST.get(f"quantidade_{rolo.nro_rolo}")
            rolo.peso = float(peso_str.replace(",", ".")) if peso_str else 0.0
            rolo.qtd_etiquetas = int(qtd_str) if qtd_str else 1
            rolos.append(rolo)
    else:
        rolos = list(rolos_queryset)

    # Para cada rolo, gera o QR Code dinâmico
    for rolo in rolos:
        if rolo.peso is None:
            rolo.peso = 0.0
        data_qr = f"{registro.materia_prima.codigo};{rolo.nro_rolo};{rolo.peso};{registro.materia_prima.localizacao};{registro.fornecedor.nome}"
        qr = qrcode.make(data_qr)
        buffer = io.BytesIO()
        qr.save(buffer, format="PNG")
        qr_code_base64 = base64.b64encode(buffer.getvalue()).decode()
        rolo.qrcode_url = f"data:image/png;base64,{qr_code_base64}"

    # Garante que o status do F045 esteja acessível
    registro.status = (
        registro.f045.status_geral if hasattr(registro, "f045") else "Aguardando F045"
    )

    # Renderiza o template PDF com os dados atualizados
    logo_url = f"file://{os.path.join(settings.STATIC_ROOT, 'img', 'logo.png')}"

    html_string = render_to_string(
        "tb050/etiqueta_lote_pdf.html",
        {
            "registro": registro,
            "rolos": rolos,
            "request": request,
            "data_atual": date.today().strftime("%d/%m/%Y"),
            "logo_url": logo_url,
        },
    )


    # Geração do PDF utilizando WeasyPrint
    with tempfile.NamedTemporaryFile(delete=True) as output:
        html = HTML(string=html_string, base_url=request.build_absolute_uri("/"))
        pdf_bytes = html.write_pdf()
        output.write(pdf_bytes)
        output.seek(0)
        response = HttpResponse(output.read(), content_type="application/pdf")
        response["Content-Disposition"] = f"filename=etiquetas_lote_{registro.id}.pdf"
        return response

from qualidade_fornecimento.models import RelacaoMateriaPrima, NormaTecnica

@login_required
def norma_aprovada(request, id):
    try:
        relacao = RelacaoMateriaPrima.objects.select_related("materia_prima").get(pk=id)
        norma_nome = relacao.materia_prima.norma

        norma = NormaTecnica.objects.get(nome_norma=norma_nome)

        if norma.aprovada:
            return JsonResponse({"aprovada": True})
        else:
            return JsonResponse({"aprovada": False, "nome": norma.nome_norma})

    except (NormaTecnica.DoesNotExist, RelacaoMateriaPrima.DoesNotExist):
        return JsonResponse({"aprovada": False})
    
from qualidade_fornecimento.models.rolo import RoloMateriaPrima
from qualidade_fornecimento.forms.rolo_forms import RoloMateriaPrimaForm

@login_required
def adicionar_rolo(request, id):
    relacao = get_object_or_404(RelacaoMateriaPrima, pk=id)

    if request.method == "POST":
        form = RoloMateriaPrimaForm(request.POST)
        if form.is_valid():
            rolo = form.save(commit=False)
            rolo.tb050 = relacao
            rolo.save()  # agora usa o save() do modelo para definir nro_rolo
            messages.success(request, f"Rolo #{rolo.nro_rolo} adicionado com sucesso!")
        else:
            messages.error(request, "Erro ao adicionar rolo. Verifique os dados.")
    return redirect("tb050_list")
